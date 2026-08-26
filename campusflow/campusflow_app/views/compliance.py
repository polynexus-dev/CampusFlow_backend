import io
import zipfile

from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from django.db import connection

from ..demo_guard import IsNotDemoTenant
from ..models.accreditation_narrative import AccreditationNarrativeDraft
from ..models.compliance import (
    ComplianceCertificateType, ComplianceCertificate,
    AccreditationCriterion, EvidenceItem, InstitutionProfile,
)
from ..models.department import Department
from ..models.finance import FinancialYear
from ..models.profile import (
    StudentProfile, TeachingStaffProfile, NonTeachingStaffProfile,
    ManagementProfile, AdministratorProfile, DepartmentHeadProfile,
)
from ..models.academics import Program
from ..serializers import (
    ComplianceCertificateTypeSerializer, ComplianceCertificateSerializer,
    AccreditationCriterionSerializer, EvidenceItemSerializer,
    AccreditationNarrativeDraftSerializer,
)
from ..permissions import IsSaaSOrCollegeAdmin, IsHMOrAbove, RequiresModule

COMPLIANCE_ADMIN_PERMS = [IsAuthenticated, IsSaaSOrCollegeAdmin, RequiresModule("compliance-center")]
from ..services.naac_ssr_export import build_naac_document
from ..services.nba_sar_export import build_nba_sar_document
from ..tasks import run_accreditation_narrative_draft


class ComplianceCertificateTypeViewSet(viewsets.ModelViewSet):
    """
    Catalog of certificate categories (UGC Recognition, Fire NOC, AICTE EOA Letter, etc.)
    that populates the upload form's category dropdown. Admin-managed.
    """
    queryset = ComplianceCertificateType.objects.all().order_by('name')
    serializer_class = ComplianceCertificateTypeSerializer
    permission_classes = COMPLIANCE_ADMIN_PERMS


class ComplianceCertificateViewSet(viewsets.ModelViewSet):
    """
    The certificate vault itself — the manual "upload document" surface for
    certificates/licenses that can only ever be supplied by the college, never
    generated from system data (UGC recognition, Trust registration, AICTE EOA
    letter, Fire Safety NOC, previous NAAC certificate, etc.).
    """
    queryset = ComplianceCertificate.objects.select_related('certificate_type', 'uploaded_by').all()
    serializer_class = ComplianceCertificateSerializer
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)


# ─────────────────────────────────────────────
# P5 — Accreditation Reporting Quick Wins
# Pure reporting over data the system already has — no new fields required.
# ─────────────────────────────────────────────

def _accreditation_xlsx_response(filename, sheets):
    """sheets: list of (sheet_title, header, rows) — one tab per section, so a
    single download covers the whole regulatory return instead of one CSV per
    table. Local to this module rather than shared with audit_portal.py's
    export helpers since these reports have a genuinely different shape
    (multi-sheet vs. one flat table per report)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_title, header, rows in sheets:
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in row])
        for column_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _accreditation_pdf_response(filename, title, subtitle, sections):
    """Same (heading, header, rows) sections shape as _accreditation_xlsx_response
    above, rendered as a submission-shaped PDF instead of a workbook — for
    reports (like the AICTE Mandatory Disclosure) that are normally submitted
    as one formatted document rather than a spreadsheet. Cell values are
    wrapped in Paragraph so long text (e.g. faculty qualifications) wraps
    within its column instead of overflowing the page."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    cell_style = styles["Normal"].clone("cell")
    cell_style.fontSize = 8
    cell_style.leading = 10

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    story = [Paragraph(title, styles["Title"]), Paragraph(subtitle, styles["Normal"]), Spacer(1, 0.5 * cm)]
    for heading, header, rows in sections:
        story.append(Paragraph(heading, styles["Heading2"]))
        col_width = doc.width / len(header)
        table_data = [[Paragraph(f"<b>{h}</b>", cell_style) for h in header]]
        for row in rows:
            table_data.append([Paragraph("" if v is None else str(v), cell_style) for v in row])
        table = Table(table_data, colWidths=[col_width] * len(header), repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.5 * cm))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def get_institution_profile():
    """Fetch (or lazily create) the tenant's InstitutionProfile — same
    get_or_create singleton idiom used by the academic calendar and clearance
    settings elsewhere in this codebase."""
    profile, _ = InstitutionProfile.objects.get_or_create(pk=1)
    return profile


class InstitutionProfileView(APIView):
    """GET/PATCH static institution-level identifiers (currently just the
    AISHE code) that regulatory returns need but no other model can supply."""
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request):
        profile = get_institution_profile()
        return Response({"aishe_code": profile.aishe_code, "updated_at": profile.updated_at})

    def patch(self, request):
        profile = get_institution_profile()
        if "aishe_code" in request.data:
            profile.aishe_code = request.data["aishe_code"]
            profile.save(update_fields=["aishe_code", "updated_at"])
        return Response({"aishe_code": profile.aishe_code, "updated_at": profile.updated_at})


class AISHEAnnualReturnView(APIView):
    """AISHE annual return: enrolment by category/gender/disability status,
    staff counts across all role types."""
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request):
        aishe_code = get_institution_profile().aishe_code
        students = StudentProfile.objects.all()
        by_category = list(students.values("category").annotate(count=Count("id")).order_by("category"))
        by_gender = list(students.values("gender").annotate(count=Count("id")).order_by("gender"))
        by_disability = list(students.values("disability_status").annotate(count=Count("id")).order_by("disability_status"))

        staff_counts = {
            "teaching": TeachingStaffProfile.objects.count(),
            "non_teaching": NonTeachingStaffProfile.objects.count(),
            "management": ManagementProfile.objects.count(),
            "administrator": AdministratorProfile.objects.count(),
            "department_head": DepartmentHeadProfile.objects.count(),
        }

        if (request.query_params.get("export") or "").lower() == "xlsx":
            return _accreditation_xlsx_response(
                "aishe_annual_return.xlsx",
                [
                    ("Institution", ["Field", "Value"], [["AISHE Code", aishe_code]]),
                    ("Enrolment by Category", ["Category", "Count"], [[r["category"], r["count"]] for r in by_category]),
                    ("Enrolment by Gender", ["Gender", "Count"], [[r["gender"], r["count"]] for r in by_gender]),
                    ("Enrolment by Disability", ["Disability Status", "Count"], [[r["disability_status"], r["count"]] for r in by_disability]),
                    ("Staff Counts", ["Role", "Count"], list(staff_counts.items())),
                ],
            )

        return Response({
            "aishe_code": aishe_code,
            "total_students": students.count(),
            "enrolment_by_category": by_category,
            "enrolment_by_gender": by_gender,
            "enrolment_by_disability_status": by_disability,
            "staff_counts": staff_counts,
            "total_staff": sum(staff_counts.values()),
            "total_departments": Department.objects.count(),
        })


class AICTEDisclosureView(APIView):
    """AICTE Mandatory Disclosure + Faculty-Student Ratio report: faculty
    qualifications/experience/designation, formatted to the disclosure template."""
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request):
        faculty = TeachingStaffProfile.objects.select_related("user", "department").all()
        faculty_rows = [
            {
                "employee_id": f.employee_id,
                "aicte_faculty_id": f.aicte_faculty_id,
                "name": f.user.get_full_name() or f.user.username,
                "department": f.department.name if f.department else None,
                "designation": f.designation,
                "aicte_cadre": f.get_aicte_cadre_display() if f.aicte_cadre else None,
                "qualifications": f.qualifications,
                "experience_years": f.experience_years,
            }
            for f in faculty
        ]

        total_students = StudentProfile.objects.count()
        total_faculty = faculty.count()
        ratio = round(total_students / total_faculty, 2) if total_faculty else None

        cadre_counts = {
            label: faculty.filter(aicte_cadre=value).count()
            for value, label in TeachingStaffProfile.CADRE_CHOICES
        }
        cadre_counts["Not Set"] = faculty.filter(aicte_cadre__isnull=True).count() + faculty.filter(aicte_cadre="").count()

        programs = Program.objects.select_related("department").filter(is_active=True)
        program_rows = [
            {
                "code": p.code, "name": p.name, "level": p.level,
                "department": p.department.name, "aicte_program_code": p.aicte_program_code,
            }
            for p in programs
        ]

        export_format = (request.query_params.get("export") or "").lower()
        disclosure_sections = [
            ("Faculty", ["Employee ID", "AICTE Faculty ID", "Name", "Department", "Designation", "AICTE Cadre", "Qualifications", "Experience (yrs)"],
             [[f["employee_id"], f["aicte_faculty_id"], f["name"], f["department"], f["designation"], f["aicte_cadre"], f["qualifications"], f["experience_years"]] for f in faculty_rows]),
            ("Programs", ["Code", "Name", "Level", "Department", "AICTE Program Code"],
             [[p["code"], p["name"], p["level"], p["department"], p["aicte_program_code"]] for p in program_rows]),
            ("Faculty-Student Ratio", ["Metric", "Value"],
             [["Total Students", total_students], ["Total Faculty", total_faculty], ["Ratio (Students:Faculty)", ratio]]),
            ("Cadre-wise Faculty Count", ["Cadre", "Count"], list(cadre_counts.items())),
        ]

        if export_format == "xlsx":
            return _accreditation_xlsx_response("aicte_mandatory_disclosure.xlsx", disclosure_sections)

        if export_format == "pdf":
            tenant_name = getattr(connection.tenant, "name", "") or connection.schema_name
            return _accreditation_pdf_response(
                "aicte_mandatory_disclosure.pdf",
                title="AICTE Mandatory Disclosure",
                subtitle=f"{tenant_name} — Faculty & Program Particulars",
                sections=disclosure_sections,
            )

        return Response({
            "faculty": faculty_rows,
            "total_students": total_students,
            "total_faculty": total_faculty,
            "student_faculty_ratio": ratio,
            "cadre_wise_faculty_count": cadre_counts,
            "programs": program_rows,
        })


class NAACExtendedProfileView(APIView):
    """NAAC Extended Profile — institution-level counts feeding the IIQA submission."""
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request):
        programs_by_level = list(Program.objects.filter(is_active=True).values("level").annotate(count=Count("id")))
        counts = {
            "departments": Department.objects.count(),
            "programs": Program.objects.filter(is_active=True).count(),
            "total_students": StudentProfile.objects.count(),
            "total_teaching_staff": TeachingStaffProfile.objects.count(),
            "total_non_teaching_staff": NonTeachingStaffProfile.objects.count(),
            "students_with_disability": StudentProfile.objects.filter(disability_status=True).count(),
        }

        if (request.query_params.get("export") or "").lower() == "xlsx":
            return _accreditation_xlsx_response(
                "naac_extended_profile.xlsx",
                [
                    ("Institution Counts", ["Metric", "Value"], list(counts.items())),
                    ("Programs by Level", ["Level", "Count"], [[r["level"], r["count"]] for r in programs_by_level]),
                ],
            )

        return Response({
            **counts,
            "programs_by_level": programs_by_level,
            "generated_at": timezone.now(),
        })


# ─────────────────────────────────────────────
# P6 — NAAC SSR/AQAR Evidence Workspace
# ─────────────────────────────────────────────

class AccreditationCriterionViewSet(viewsets.ModelViewSet):
    """The NAAC/NBA criteria catalog (Admin/IQAC managed)."""
    queryset = AccreditationCriterion.objects.all()
    serializer_class = AccreditationCriterionSerializer
    permission_classes = COMPLIANCE_ADMIN_PERMS + [IsNotDemoTenant]


class EvidenceItemViewSet(viewsets.ModelViewSet):
    """
    Draft -> Submitted -> Signed Off workflow: Department Heads/Faculty attach
    evidence (a file, or a pointer at an existing record / Phase-1 certificate
    instead of re-uploading), an IQAC reviewer (Department Head or above)
    signs off per criterion.
    """
    queryset = EvidenceItem.objects.select_related(
        "criterion", "department", "financial_year", "uploaded_by", "signed_off_by", "linked_certificate",
    ).all()
    serializer_class = EvidenceItemSerializer
    permission_classes = [IsAuthenticated, IsNotDemoTenant, RequiresModule("compliance-center")]

    def get_queryset(self):
        qs = super().get_queryset()
        criterion = self.request.query_params.get("criterion")
        financial_year = self.request.query_params.get("financial_year")
        department = self.request.query_params.get("department")
        if criterion:
            qs = qs.filter(criterion_id=criterion)
        if financial_year:
            qs = qs.filter(financial_year_id=financial_year)
        if department:
            qs = qs.filter(department_id=department)
        return qs

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)


class SubmitEvidenceItemView(APIView):
    """POST /api/evidence-items/<int:pk>/submit/ — Draft -> Submitted."""
    permission_classes = [IsAuthenticated, IsNotDemoTenant, RequiresModule("compliance-center")]

    def post(self, request, pk):
        try:
            item = EvidenceItem.objects.get(pk=pk)
        except EvidenceItem.DoesNotExist:
            return Response({"error": "Evidence item not found."}, status=status.HTTP_404_NOT_FOUND)
        if item.status != EvidenceItem.STATUS_DRAFT:
            return Response({"error": f"Cannot submit an item that is already {item.status}."}, status=status.HTTP_400_BAD_REQUEST)
        item.submit()
        return Response({"message": "Evidence submitted for IQAC review.", "evidence": EvidenceItemSerializer(item).data})


class SignOffEvidenceItemView(APIView):
    """POST /api/evidence-items/<int:pk>/sign-off/ — Submitted -> Signed Off.
    Restricted to Department Head or above, the same bar as other HM-level
    re-approval actions in this codebase (see IsHMOrAbove)."""
    permission_classes = [IsAuthenticated, IsHMOrAbove, IsNotDemoTenant, RequiresModule("compliance-center")]

    def post(self, request, pk):
        try:
            item = EvidenceItem.objects.get(pk=pk)
        except EvidenceItem.DoesNotExist:
            return Response({"error": "Evidence item not found."}, status=status.HTTP_404_NOT_FOUND)
        if item.status != EvidenceItem.STATUS_SUBMITTED:
            return Response({"error": "Only submitted evidence can be signed off."}, status=status.HTTP_400_BAD_REQUEST)
        item.sign_off(request.user)
        return Response({"message": "Evidence signed off.", "evidence": EvidenceItemSerializer(item).data})


# ─────────────────────────────────────────────
# P7 — AI-Assisted Narrative Drafting (see campusflow_app/ai_narrative.py,
# models/accreditation_narrative.py, run_accreditation_narrative_draft task)
# ─────────────────────────────────────────────

class CriterionNarrativeDraftRequestView(APIView):
    """
    POST /accreditation-criteria/<int:pk>/narrative-draft/?financial_year=<id>
    Queues an AI-drafted first-pass SSR/AQAR narrative for one criterion+year.
    Same permission bar as the rest of this file's admin/reporting views —
    not the broader IsHMOrAbove used only for evidence sign-off.
    """
    permission_classes = COMPLIANCE_ADMIN_PERMS + [IsNotDemoTenant]

    def post(self, request, pk):
        try:
            criterion = AccreditationCriterion.objects.get(pk=pk)
        except AccreditationCriterion.DoesNotExist:
            return Response({"error": "Accreditation criterion not found."}, status=status.HTTP_404_NOT_FOUND)

        financial_year_id = request.query_params.get("financial_year")
        if not financial_year_id:
            return Response({"error": "financial_year query param is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            financial_year = FinancialYear.objects.get(pk=financial_year_id)
        except FinancialYear.DoesNotExist:
            return Response({"error": "Financial year not found."}, status=status.HTTP_404_NOT_FOUND)

        if not criterion.evidence.filter(financial_year=financial_year).exists():
            return Response(
                {"error": "No evidence items recorded for this criterion and financial year yet — "
                          "add evidence before requesting an AI draft."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # A fresh request supersedes any draft still awaiting review for the same criterion+year.
        AccreditationNarrativeDraft.objects.filter(
            criterion=criterion, financial_year=financial_year, status="pending",
        ).update(status="rejected")

        draft = AccreditationNarrativeDraft.objects.create(
            criterion=criterion, financial_year=financial_year, requested_by=request.user,
        )
        run_accreditation_narrative_draft.delay(connection.schema_name, draft.id)

        return Response(
            AccreditationNarrativeDraftSerializer(draft).data,
            status=status.HTTP_202_ACCEPTED,
        )


class AccreditationNarrativeDraftViewSet(viewsets.ModelViewSet):
    """
    List/retrieve drafts (for polling a pending one, and browsing history),
    and PATCH to edit narrative_text/caveats before applying — every other
    field is read-only (see AccreditationNarrativeDraftSerializer).
    """
    queryset = AccreditationNarrativeDraft.objects.select_related(
        "criterion", "financial_year", "requested_by", "applied_by",
    ).all()
    serializer_class = AccreditationNarrativeDraftSerializer
    permission_classes = COMPLIANCE_ADMIN_PERMS + [IsNotDemoTenant]

    def get_queryset(self):
        qs = super().get_queryset()
        criterion = self.request.query_params.get("criterion")
        financial_year = self.request.query_params.get("financial_year")
        if criterion:
            qs = qs.filter(criterion_id=criterion)
        if financial_year:
            qs = qs.filter(financial_year_id=financial_year)
        return qs


class NarrativeDraftApplyView(APIView):
    """POST /narrative-drafts/<int:pk>/apply/ — marks a draft as the current narrative for its criterion+year."""
    permission_classes = COMPLIANCE_ADMIN_PERMS + [IsNotDemoTenant]

    def post(self, request, pk):
        try:
            draft = AccreditationNarrativeDraft.objects.get(pk=pk)
        except AccreditationNarrativeDraft.DoesNotExist:
            return Response({"error": "Narrative draft not found."}, status=status.HTTP_404_NOT_FOUND)
        if draft.status != "pending":
            return Response(
                {"error": f"This draft is already {draft.status} and cannot be applied."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        draft.status = "applied"
        draft.applied_at = timezone.now()
        draft.applied_by = request.user
        draft.save(update_fields=["status", "applied_at", "applied_by"])

        return Response(AccreditationNarrativeDraftSerializer(draft).data)


class NarrativeDraftRejectView(APIView):
    """POST /narrative-drafts/<int:pk>/reject/"""
    permission_classes = COMPLIANCE_ADMIN_PERMS + [IsNotDemoTenant]

    def post(self, request, pk):
        try:
            draft = AccreditationNarrativeDraft.objects.get(pk=pk)
        except AccreditationNarrativeDraft.DoesNotExist:
            return Response({"error": "Narrative draft not found."}, status=status.HTTP_404_NOT_FOUND)
        if draft.status != "pending":
            return Response(
                {"error": f"This draft is already {draft.status}."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        draft.status = "rejected"
        draft.save(update_fields=["status"])
        return Response(AccreditationNarrativeDraftSerializer(draft).data)


class SSRExportView(APIView):
    """
    Admin export of the full evidence tree as the SSR/AQAR submission package:
    a JSON tree (criterion -> evidence items) for the manifest, plus a ZIP of
    every evidence file, grouped by criterion code.
    """
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request):
        financial_year = request.query_params.get("financial_year")
        export_format = (request.query_params.get("export") or "").lower()

        if export_format == "docx":
            financial_year_obj = None
            if financial_year:
                financial_year_obj = FinancialYear.objects.filter(pk=financial_year).first()
                if not financial_year_obj:
                    return Response({"error": "Financial year not found."}, status=status.HTTP_404_NOT_FOUND)
            buffer = build_naac_document(financial_year_obj)
            filename = f"naac_aqar_{financial_year_obj.label}.docx" if financial_year_obj else "naac_ssr.docx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        criteria = AccreditationCriterion.objects.prefetch_related("evidence").all()

        tree = []
        for criterion in criteria:
            items = criterion.evidence.all()
            if financial_year:
                items = items.filter(financial_year_id=financial_year)
            entry = {
                "criterion": AccreditationCriterionSerializer(criterion).data,
                "evidence": EvidenceItemSerializer(items, many=True).data,
            }
            if financial_year:
                # Narratives are year-scoped, so only attach one when the
                # export itself is scoped to a single financial year.
                applied_narrative = criterion.narrative_drafts.filter(
                    financial_year_id=financial_year, status="applied",
                ).order_by("-applied_at").first()
                entry["narrative"] = applied_narrative.narrative_text if applied_narrative else None
            tree.append(entry)

        if export_format == "zip":
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for criterion in criteria:
                    items = criterion.evidence.all()
                    if financial_year:
                        items = items.filter(financial_year_id=financial_year)
                    for item in items:
                        if item.file:
                            try:
                                zf.writestr(f"{criterion.code}/{item.id}_{item.file.name.split('/')[-1]}", item.file.read())
                            except Exception:
                                continue
            response = HttpResponse(buffer.getvalue(), content_type="application/zip")
            response["Content-Disposition"] = 'attachment; filename="ssr_aqar_evidence_package.zip"'
            return response

        return Response({"criteria": tree})


# ─────────────────────────────────────────────
# P8 — NBA SAR Document Generator
# ─────────────────────────────────────────────

class NBASARExportView(APIView):
    """
    GET /api/academics/programs/<int:program_id>/nba-sar-export/?financial_year=<id>
    Compiles the program's NBA-tagged criteria (applied narrative + evidence,
    same source data as SSRExportView above), its CO-PO articulation matrix,
    and its PO/PSO/PEO attainment (services/outcome_attainment.py — the same
    computation ProgramOutcomeAttainmentView exposes as CSV) into a single
    submission-shaped .docx. Same permission bar as SSRExportView: this is an
    institution-level regulatory export, not a routine faculty-facing report.
    """
    permission_classes = COMPLIANCE_ADMIN_PERMS

    def get(self, request, program_id):
        program = Program.objects.select_related("department").filter(pk=program_id).first()
        if not program:
            return Response({"error": "Program not found."}, status=status.HTTP_404_NOT_FOUND)

        financial_year = None
        financial_year_id = request.query_params.get("financial_year")
        if financial_year_id:
            financial_year = FinancialYear.objects.filter(pk=financial_year_id).first()
            if not financial_year:
                return Response({"error": "Financial year not found."}, status=status.HTTP_404_NOT_FOUND)

        buffer = build_nba_sar_document(program, financial_year)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        response["Content-Disposition"] = f'attachment; filename="nba_sar_{program.code}.docx"'
        return response
